import mysql.connector
import csv
import xml.dom.minidom as md
import xml.etree.ElementTree as ET
from datetime import datetime
import os
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl

# Conectar ao Data Warehouse MySQL
def conectar_bd():
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="Litha003&",  # Ajuste com sua senha
            database="DW_pi"  # Conectando ao Data Warehouse
        )
        return conn
    except mysql.connector.Error as err:
        print(f"Erro ao conectar ao MySQL: {err}")
        exit(1)

# Função para buscar dados do aluno no Data Warehouse
def buscar_dados_aluno(ra):
    conn = conectar_bd()
    cursor = conn.cursor()
    
    query = """
    SELECT 
        a.nome AS aluno_nome,
        d.nome AS disciplina,
        t.data_aula,
        f.presente,
        a.curso,
        a.email,
        prof.nome AS professor_nome,
        turma.nome AS turma_nome,
        loc.sala,
        loc.bloco,
        loc.campus
    FROM FatoPresenca f
    JOIN DimAluno a ON a.id_aluno = f.id_aluno
    JOIN DimDisciplina d ON d.iddisciplina = f.id_disciplina
    JOIN DimTempo t ON t.id_tempo = f.id_tempo
    JOIN DimProfessor prof ON prof.id_professor = f.id_professor
    JOIN DimTurma turma ON turma.idturma = f.id_turma
    JOIN DimLocalizacao loc ON loc.id_localizacao = f.id_localizacao
    WHERE a.ra = %s
    ORDER BY t.data_aula DESC;
    """
    
    cursor.execute(query, (ra,))
    resultados = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return resultados

# Função para calcular estatísticas de presença
def calcular_estatisticas(resultados):
    if not resultados:
        return {}
    
    total_aulas = len(resultados)
    presencas = sum(1 for row in resultados if row[3])  # row[3] é o campo 'presente'
    ausencias = total_aulas - presencas
    percentual_presenca = (presencas / total_aulas) * 100 if total_aulas > 0 else 0
    
    return {
        'total_aulas': total_aulas,
        'presencas': presencas,
        'ausencias': ausencias,
        'percentual_presenca': round(percentual_presenca, 2)
    }

# Função para gerar relatório CSV
def gerar_csv(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.csv"
    
    with open(nome_arquivo, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        
        # Cabeçalho do relatório
        writer.writerow(['=== RELATÓRIO DE PRESENÇA - DATA WAREHOUSE ==='])
        writer.writerow([f'Aluno: {aluno_nome}'])
        writer.writerow([f'RA: {ra}'])
        writer.writerow([f'Curso: {resultados[0][4]}'])
        writer.writerow([f'Email: {resultados[0][5]}'])
        writer.writerow([''])
        
        # Estatísticas
        stats = calcular_estatisticas(resultados)
        writer.writerow(['=== ESTATÍSTICAS ==='])
        writer.writerow([f'Total de Aulas: {stats["total_aulas"]}'])
        writer.writerow([f'Presenças: {stats["presencas"]}'])
        writer.writerow([f'Ausências: {stats["ausencias"]}'])
        writer.writerow([f'Percentual de Presença: {stats["percentual_presenca"]}%'])
        writer.writerow([''])
        
        # Cabeçalho da tabela de dados
        writer.writerow(['Aluno', 'Disciplina', 'Data da Aula', 'Presença', 'Professor', 'Turma', 'Sala', 'Bloco', 'Campus'])
        
        for row in resultados:
            aluno_nome, disciplina, data_aula, presente, curso, email, professor, turma, sala, bloco, campus = row
            status = "Presente" if presente else "Ausente"
            writer.writerow([aluno_nome, disciplina, data_aula.strftime("%Y-%m-%d"), status, professor, turma, sala, bloco, campus])
    
    print(f" Relatório CSV gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para gerar relatório XML
def gerar_xml(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.xml"
    
    # Criar a estrutura XML
    root = ET.Element("RelatorioPresenca")
    root.set("gerado_em", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # Informações do aluno
    aluno_element = ET.SubElement(root, "Aluno")
    aluno_element.set("RA", str(ra))
    aluno_element.set("Nome", aluno_nome)
    aluno_element.set("Curso", resultados[0][4])
    aluno_element.set("Email", resultados[0][5])
    
    # Estatísticas
    stats = calcular_estatisticas(resultados)
    estatisticas = ET.SubElement(aluno_element, "Estatisticas")
    ET.SubElement(estatisticas, "TotalAulas").text = str(stats["total_aulas"])
    ET.SubElement(estatisticas, "Presencas").text = str(stats["presencas"])
    ET.SubElement(estatisticas, "Ausencias").text = str(stats["ausencias"])
    ET.SubElement(estatisticas, "PercentualPresenca").text = str(stats["percentual_presenca"])
    
    # Registros de presença
    registros = ET.SubElement(aluno_element, "Registros")
    
    for row in resultados:
        aluno_nome, disciplina, data_aula, presente, curso, email, professor, turma, sala, bloco, campus = row
        status = "Presente" if presente else "Ausente"
        
        presenca = ET.SubElement(registros, "Registro")
        ET.SubElement(presenca, "Disciplina").text = disciplina
        ET.SubElement(presenca, "DataAula").text = data_aula.strftime("%Y-%m-%d")
        ET.SubElement(presenca, "Status").text = status
        ET.SubElement(presenca, "Professor").text = professor
        ET.SubElement(presenca, "Turma").text = turma
        
        localizacao = ET.SubElement(presenca, "Localizacao")
        ET.SubElement(localizacao, "Sala").text = sala
        ET.SubElement(localizacao, "Bloco").text = bloco
        ET.SubElement(localizacao, "Campus").text = campus
    
    # Salvar o XML formatado
    xmlstr = ET.tostring(root, encoding='utf-8')
    dom = md.parseString(xmlstr)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    with open(nome_arquivo, "w", encoding='utf-8') as f:
        f.write(pretty_xml)
    
    print(f" Relatório XML gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para gerar relatório PDF
def gerar_pdf(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.pdf"
    
    # Criar PDF
    doc = SimpleDocTemplate(nome_arquivo, pagesize=letter)
    elements = []
    
    # Estilos
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']
    
    # Título
    elements.append(Paragraph(f"Relatório de Presença - Data Warehouse", title_style))
    elements.append(Paragraph(f"Aluno: {aluno_nome} (RA: {ra})", subtitle_style))
    elements.append(Paragraph(f"Curso: {resultados[0][4]}", normal_style))
    elements.append(Paragraph(f"Email: {resultados[0][5]}", normal_style))
    elements.append(Paragraph("<br/><br/>", normal_style))
    
    # Estatísticas
    stats = calcular_estatisticas(resultados)
    elements.append(Paragraph("Estatísticas de Presença:", subtitle_style))
    elements.append(Paragraph(f"Total de Aulas: {stats['total_aulas']}", normal_style))
    elements.append(Paragraph(f"Presenças: {stats['presencas']}", normal_style))
    elements.append(Paragraph(f"Ausências: {stats['ausencias']}", normal_style))
    elements.append(Paragraph(f"Percentual de Presença: {stats['percentual_presenca']}%", normal_style))
    elements.append(Paragraph("<br/><br/>", normal_style))
    
    # Dados para a tabela
    data = [['Disciplina', 'Data', 'Presença', 'Professor', 'Turma', 'Local']]
    
    for row in resultados:
        aluno_nome, disciplina, data_aula, presente, curso, email, professor, turma, sala, bloco, campus = row
        status = "Presente" if presente else "Ausente"
        local = f"{sala}/{bloco}"
        data.append([disciplina, data_aula.strftime("%Y-%m-%d"), status, professor, turma, local])
    
    # Criar tabela
    table = Table(data)
    
    # Estilo da tabela
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ])
    
    table.setStyle(style)
    elements.append(table)
    
    # Construir o PDF
    doc.build(elements)
    
    print(f" Relatório PDF gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para gerar relatório Excel
def gerar_excel(ra, resultados):
    if not resultados:
        return False
        
    aluno_nome = resultados[0][0]
    nome_arquivo = f"relatorio_presenca_RA_{ra}.xlsx"
    
    # Criar uma nova planilha
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Presença RA {ra}"
    
    # Informações do cabeçalho
    ws.append([f'RELATÓRIO DE PRESENÇA - DATA WAREHOUSE'])
    ws.append([f'Aluno: {aluno_nome}'])
    ws.append([f'RA: {ra}'])
    ws.append([f'Curso: {resultados[0][4]}'])
    ws.append([f'Email: {resultados[0][5]}'])
    ws.append([f'Gerado em: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    ws.append([''])
    
    # Estatísticas
    stats = calcular_estatisticas(resultados)
    ws.append(['ESTATÍSTICAS DE PRESENÇA'])
    ws.append([f'Total de Aulas: {stats["total_aulas"]}'])
    ws.append([f'Presenças: {stats["presencas"]}'])
    ws.append([f'Ausências: {stats["ausencias"]}'])
    ws.append([f'Percentual de Presença: {stats["percentual_presenca"]}%'])
    ws.append([''])
    
    # Cabeçalho da tabela de dados
    header_row = ws.max_row + 1
    ws.append(['Disciplina', 'Data da Aula', 'Presença', 'Professor', 'Turma', 'Sala', 'Bloco', 'Campus'])
    
    # Estilizar cabeçalho da tabela
    for cell in ws[header_row]:
        cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        cell.fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    # Adicionar dados
    for row in resultados:
        aluno_nome, disciplina, data_aula, presente, curso, email, professor, turma, sala, bloco, campus = row
        status = "Presente" if presente else "Ausente"
        ws.append([disciplina, data_aula.strftime("%Y-%m-%d"), status, professor, turma, sala, bloco, campus])
        
        # Colorir linha baseada na presença
        current_row = ws.max_row
        if not presente:  # Ausente
            for cell in ws[current_row]:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min((max_length + 2), 50)  # Limitar largura máxima
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar o arquivo
    wb.save(nome_arquivo)
    
    print(f" Relatório Excel gerado com sucesso: {nome_arquivo}")
    return nome_arquivo

# Função para listar alunos disponíveis (para ajudar o usuário)
def listar_alunos_disponiveis():
    conn = conectar_bd()
    cursor = conn.cursor()
    
    query = """
    SELECT DISTINCT a.ra, a.nome, a.curso
    FROM DimAluno a
    JOIN FatoPresenca f ON a.id_aluno = f.id_aluno
    ORDER BY a.nome
    LIMIT 10;
    """
    
    try:
        cursor.execute(query)
        resultados = cursor.fetchall()
        
        if resultados:
            print("\n Alguns alunos disponíveis no sistema:")
            print("RA\t\tNome\t\t\tCurso")
            print("-" * 60)
            for ra, nome, curso in resultados:
                print(f"{ra}\t\t{nome[:20]:<20}\t{curso}")
        
    except mysql.connector.Error as err:
        print(f"Erro ao listar alunos: {err}")
    finally:
        cursor.close()
        conn.close()

# Função principal
def main():
    print("=" * 60)
    print("    SISTEMA DE RELATÓRIOS - DATA WAREHOUSE DW_pi")
    print("=" * 60)
    
    # Verificar conexão
    try:
        conn = conectar_bd()
        conn.close()
        print(" Conexão com o Data Warehouse estabelecida com sucesso!")
    except:
        print(" Erro ao conectar com o Data Warehouse!")
        return
    
    # Listar alguns alunos disponíveis
    listar_alunos_disponiveis()
    
    # Solicitar o RA do aluno
    ra_input = input("\nDigite o RA do aluno: ")
    
    # Buscar dados
    print(" Buscando dados no Data Warehouse...")
    resultados = buscar_dados_aluno(ra_input)
    
    if not resultados:
        print(" Nenhum registro encontrado para este RA no Data Warehouse.")
        return
    
    print(f" Encontrados {len(resultados)} registros de presença!")
    
    continuar_gerando = True
    
    while continuar_gerando:
        # Menu de opções
        print("\n" + "=" * 40)
        print("Escolha o formato do relatório:")
        print("1 - CSV (planilha simples)")
        print("2 - XML (dados estruturados)")
        print("3 - PDF (documento formatado)")
        print("4 - Excel (planilha avançada)")
        print("=" * 40)
        
        opcao = input("Digite o número da opção desejada: ")
        
        arquivo_gerado = None
        
        print(" Gerando relatório...")
        
        if opcao == "1":
            arquivo_gerado = gerar_csv(ra_input, resultados)
        elif opcao == "2":
            arquivo_gerado = gerar_xml(ra_input, resultados)
        elif opcao == "3":
            arquivo_gerado = gerar_pdf(ra_input, resultados)
        elif opcao == "4":
            arquivo_gerado = gerar_excel(ra_input, resultados)
        else:
            print(" Opção inválida!")
            continue
        
        if arquivo_gerado:
            # Perguntar se deseja gerar outro tipo de relatório
            resposta = input("\nDeseja gerar outro tipo de relatório para o mesmo RA? (s/n): ").lower()
            continuar_gerando = resposta == 's' or resposta == 'sim'
        else:
            continuar_gerando = False
    
    print("\n Programa finalizado. Obrigado por utilizar o sistema de relatórios!")

# Executar o programa
if __name__ == "__main__":
    # Verificar se as bibliotecas necessárias estão instaladas
    print("Verificando dependências...")
    try:
        import mysql.connector
        import openpyxl
        import reportlab
        print(" Todas as dependências estão instaladas!")
        print("Se alguma estiver faltando, instale usando:")
        print("pip install mysql-connector-python openpyxl reportlab")
        print()
    except ImportError as e:
        print(f" Dependência faltando: {e}")
        print("Execute: pip install mysql-connector-python openpyxl reportlab")
        exit(1)
    
    main()